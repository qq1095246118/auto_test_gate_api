# -*- coding:utf-8 -*-
"""
@FileName: data_helper.py
@Desc    :   文件处理模块
"""
import json
import os

import allure
import openpyxl
import pandas as pd
import xlrd
from openpyxl import load_workbook

from common.api_key import ApiKey
from utility.client import request


def read_file(filepath):
    """读取json文件"""
    with open(filepath, 'r', encoding='UTF-8') as file:
        content = file.read()
    return json.loads(content)


# 根目录
cur_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# 运行文件目录
xls_path = os.path.join(cur_path, 'testdata', 'xlsx', 'test_case.xlsx')


def read_excel_file(file_path):
    """
        Excel数据读取，默认返回每行数据，以表头为key,内容为value
    :return: 以list返回，每行数据为一个dict
    """
    excel_data = []
    # 将excel进行实例化
    book = xlrd.open_workbook(file_path)
    # 通过下标方法读取sheet值
    sheet = book.sheet_by_index(0)
    # 遍历获取Excel表信息,从第一行开始读取,根据列数读取数据
    for data in [dict(zip(sheet.row_values(0), sheet.row_values(row))) for row in range(1, sheet.nrows)]:
        excel_data.append(data)
    return excel_data


def excel_data_handle(xls_paths):
    """
        Excel数据处理，将运行结果写入Excel
    """
    data_list = []
    # 将excel进行实例化
    book = xlrd.open_workbook(xls_path)
    # 通过下标方法读取sheet值
    sheet1 = book.sheet_by_index(0)
    # 读取Excel文件
    excel = openpyxl.load_workbook(xls_path)
    # 读取所有sheet页
    sheets = excel.sheetnames
    # for循环遍历当前sheet页
    for sheet in sheets:
        sheet_temp = excel[sheet]
        row = 0
        for value in list(sheet_temp.values):
            # 只有表格中第一行有值才能进入
            if type(value[0]) is str:
                dict1 = {}
                # 获取第一行表头数据
                list1 = sheet1.row_values(0)
                # 遍历获取api参数，并放入list
                list2 = []
                for i in range(len(value)):
                    list2.append(value[i])
                row += 1
                count = 0
                for i in list2:
                    if i:
                        try:
                            dict1['row'] = row
                            dict1['{}'.format(list1[count])] = eval(i)
                        except:
                            dict1['{}'.format(list1[count])] = i
                    count += 1
                if dict1.get('title') == 'title':
                    pass
                else:
                    data_list.append(dict1)
    return data_list


def requests_result(params):
    for param_data in params:
        # allure.dynamic.title(param_data['title'])
        print(f"接口名称:-------------->{param_data['title']}")
        print(f"请求方式:-------------->{param_data['method']}")
        # print(f"请求地址:-------------->{param_data['url']}")
        # print(f"请求JSON:------------->\n{param_data['body']}")
        # print(f"正在将参数映射到接口类:------------>ApiKey")
        # getattr(ApiKey, param_data[3])(**param_data)
        if param_data['method'].lower() == "post":
            print(param_data)
            if 'user' in str(param_data):
                # 封装请求
                res = request.post(param_data['url'], param_data['body'], param_data['user'])
            else:
                res = request.post(param_data['url'], param_data['body'])
        else:
            res = request.get(param_data['url'])
        print(f"正在将结果写入Excel:------------>")
        result_data_write(xls_path, param_data, res)
        print(f"已完成写入:------------>")

def result_data_write(xls_path, data, response_json):
    """
        写入执行结果
    """
    # 读取Excel文件
    excel = openpyxl.load_workbook(xls_path)
    # 读取所有sheet页
    sheet = excel.sheetnames[0]
    # for循环遍历当前sheet页
    # 遍历读取当前Excel中所有sheet中内容
    sheet_temp = excel[sheet]
    if response_json != [] and "message" not in response_json:
        sheet_temp.cell(row=int(data['row']), column=5).value = str(response_json)
        sheet_temp.cell(row=data['row'], column=6).value = 'PASSED'

    # 执行Excel的保存
    excel.save(xls_path)

if __name__ == '__main__':
    data = excel_data_handle(xls_path)
    # # print(data)
    a = requests_result(data)
    # data2 = excel_data_handle(xls_path)

    # print(data2)
