# -*- coding: utf-8 -*-
"""
@File    :   excel_request.py
@Time    :   2022/11/30/17:33
@Author  :   Yan-QC
@Desc    :   Excel数据驱动模块
"""

import hashlib
import os
import traceback
from time import time

import openpyxl
import xlrd
import requests
import json
import pytest
from utility.tools import *

# 获取项目根路径
cur_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
# 运行文件目录
xls_paths = os.path.join(cur_path, 'other', 'test_case_xls')


def Read_Excel(excel_name):
    print(f'正在获取{excel_name}Excel数据------------>')
    data_list = []
    excel_dict = {}
    # 文件路径拼装
    xls_path = os.path.join(xls_paths, excel_name)
    # 将excel进行实例化
    book = xlrd.open_workbook(xls_path)
    # 通过下标方法读取sheet值
    sheet = book.sheet_by_index(0)
    print(f'即将执行{excel_name},Excel数据格式处理正在运行------------>')
    # 将Excel中的数据去除换行符与空格  .replace("\n ", ""), 循环读取每行数据
    for data in [dict(zip(sheet.row_values(0), sheet.row_values(row))) for row in range(1, sheet.nrows)]:
        for key, value in data.items():
            # body中包含sign故作特殊处理
            if key == 'body':
                excel_dict[key] = value.replace("\n", "")
            else:
                excel_dict[key] = value.replace("\n ", "")
        data_list.append(excel_dict)
    return data_list


def excel_data_handle(excel_name):
    """
        Excel数据处理，将运行结果写入Excel
    """
    print(f'正在获取{excel_name}Excel数据------------>')
    data_list = []
    # 文件路径拼装
    xls_path = os.path.join(xls_paths, excel_name)

    # 将excel进行实例化
    book = xlrd.open_workbook(xls_path)
    # 通过下标方法读取sheet值
    sheet1 = book.sheet_by_index(0)

    # 读取Excel文件
    excel = openpyxl.load_workbook(xls_path)
    # 读取所有sheet页
    sheets = excel.sheetnames
    print(f'即将执行{excel_name},Excel数据格式处理正在运行------------>')
    # for循环遍历当前sheet页
    for sheet in sheets:
        sheet_temp = excel[sheet]
        row = 0
        for value in list(sheet_temp.values):
            # 只有表格中第一行有值才能进入
            if type(value[0]) is str:
                dict1 = {}
                # 获取第一行表头数据
                list1 = sheet1.row_values(0)
                # 遍历获取api参数，并放入list
                list2 = []
                for i in range(len(value)):
                    list2.append(value[i])
                row += 1
                count = 0
                for i in list2:
                    if i:
                        try:
                            if 'update' in list1[count]:
                                dict1['{}'.format(list1[count])] = f"{eval(i)}"
                            else:
                                dict1['row'] = row
                                dict1['{}'.format(list1[count])] = eval(i)
                        except:
                            if 'update' in list1[count]:
                                dict1['{}'.format(list1[count])] = f"{i}"
                            else:
                                dict1['{}'.format(list1[count])] = i
                    count += 1
                if dict1.get('title') == 'title':
                    pass
                else:
                    data_list.append(dict1)
    print(f'执行{excel_name},Excel数据格式处理成功------------>')
    return data_list


def result_data_write(excel_name, item, response_json=None, status=True, error_info=None):
    """
        写入执行结果
    """
    # 文件路径拼装
    xls_path = os.path.join(cur_path, 'other', 'test_case_xls', excel_name)
    # 读取Excel文件
    excel = openpyxl.load_workbook(xls_path)
    # 读取所有sheet页
    sheet = excel.sheetnames[0]
    # for循环遍历当前sheet页
    # 遍历读取当前Excel中所有sheet中内容
    sheet_temp = excel[sheet]
    if status:
        # 判断预期与实际结果，若内容不相同则写入response,若预期结果一致写入PASSED并，写入结果
        # if 'assert_field' in item and 'expect' in item:
        if response_json['errorMsg'] == '成功' and response_json['errorCode'] == 'SUCCESS':
            if item.get('assert_field') is not None and item.get('expect') is not None:
                if item['expect'] == response_json.get(item['assert_field']):
                    sheet_temp.cell(row=item['row'], column=10).value = 'PASSED'
                else:
                    sheet_temp.cell(row=item['row'], column=10).value = 'FAILED'
            else:
                sheet_temp.cell(row=item['row'], column=10).value = 'PASSED'
            sheet_temp.cell(row=item['row'], column=8).value = str(response_json)
        else:
            sheet_temp.cell(row=item['row'], column=10).value = 'FAILED'
            sheet_temp.cell(row=item['row'], column=8).value = str(response_json)
    else:
        if error_info is None:
            error_info = '接口运行发生错误，请检查当前请求body与用例信息!'
        sheet_temp.cell(row=item['row'], column=10).value = 'FAILED'
        sheet_temp.cell(row=item['row'], column=8).value = str(response_json)
    # assert False, f"请检查文件信息"
    # 执行Excel的保存
    excel.save(xls_path)


def Alone_ExcelCase(row):
    # 获取项目根路径
    cur_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    # 文件路径拼装
    xls_path = os.path.join(cur_path, 'other', 'test_case_xls')
    # 读取指定文件夹下的文件信息
    xls_names = os.listdir(xls_path)
    # 将excel进行实例化
    book = xlrd.open_workbook(xls_path)
    # 通过name值进行读取sheet
    sheet = book.sheet_by_name('Sheet1')
    # 读取指定行数
    return dict(zip(sheet.row_values(0), sheet.row_values(row)))


def Requests_result(item):
    print(f"正在请求接口-------------->{item['title']}")
    print(f"正在请求接口URL:---------->{item['url']}")
    print(f"请求接口JSON:------------>{item['body']}")
    if 'update' in item and item['row'] > 1:
        print(f"检测到更新字段,正在替换json内字段:------------>{item['update']}")
        print(get_value('result'))
    # BPP token写死，替换所有bpp请求中的token与sign
    if 'bpp' in item['url']:
        token = 'store:2005223350b0aa5bdc63eb21b0'
        sign_md5 = token + str(round(time() * 1000)) + "key=" + "MKnEu6zaS04N23XoMUL8GOwOKIQwXMvT"
        bpp_sign = hashlib.md5(sign_md5.encode('utf-8')).hexdigest().upper()
        # 请求body中携带token，取默认赋值
        if 'token' in str(item['body']):
            item['token'] = token
            item['sign'] = bpp_sign
    # 封装请求
    res = requests.request(method=item['method'],
                           url=item['url'],
                           json=item['body'],
                           headers=item['header'])
    # 返回结果格式处理，解析一般json，若无法解析json则返回原生response
    try:
        res = res.json()
    except ValueError:
        return res
    if isinstance(res, bool):
        return res
    return res


def response_data(excel_name, items):
    # 遍历运行Excel表中所有接口信息,成功PASSED,失败FAILED
    for item in items:
        try:
            response_result = Requests_result(item)
            print(f'正在请求接口成功返回结果为:------------>{response_result}')
            # 将返回结果写入Excel
            if 'errorCode' in response_result:
                # 断言请求返回的结果errorCode是否为True
                if response_result['errorCode'] == 'SUCCESS':
                    # 存放上个结果result，用作下个接口取值使用
                    set_value('result', response_result['data'])
                    # 对预期结果与实际结果进行比较,若与预期结果不一致则将单条运行结果FAILED写入Excel
                    result_data_write(excel_name, item, response_result)
                else:
                    # 执行或获取接口参数失败，将单条运行结果：FAILED
                    result_data_write(excel_name, item, response_result, False)
        except Exception:
            print(f'接口运行异常报错,正在打印报错堆栈信息---------->{str(traceback.print_exc())}')
            # 若运行结果报错则将单条接口Failed写入Excel
            result_data_write(excel_name, item, status=False)
            assert False, f"用例运行失败，请检查{excel_name}内用例！"


class TestReadExcelApi:
    # 通过参数化的方式进行导入excel的数据
    # @pytest.mark.parametrize('excel_name', os.listdir(xls_paths))
    @pytest.mark.parametrize('excel_name', ['test.he_case.xlsx'])
    def test_run_excel_api(self, excel_name):
        set_value('phone', 15258822463)
        items = excel_data_handle(excel_name)
        response = response_data(excel_name, items)

    # @pytest.mark.parametrize('excel_name', os.listdir(xls_paths))
    # @pytest.mark.parametrize('excel_name', ['test_vms.xlsx'])
    def test_001(self, excel_name):
        set_value('phone', 15258822463)
        data_list = excel_data_handle(excel_name)
        print('####' * 20)
        print(data_list)
        # for item in data_list:
        #     result_data_write(excel_name, item)

    def test_02(self):
        set_value('phone', 15258822463)
        print(get_value('phone'))
        case = Alone_ExcelCase(3)
        print(case)
        response_result = Requests_result(case)
        print('*****' * 50)
        print(response_result)
        # print('*****' *50)
        # print(case)
        # print(response_result)
        # 断言请求返回的结果是否为True
        assert response_result is True, f"预期结果：True 实际结果：{response_result}"


if __name__ == '__main__':
    pytest.main(['-vs'])
